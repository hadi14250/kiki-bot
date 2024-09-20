import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

def addInvalidUsersToExcelSheet(user):
    excel_file_path = "Daily_User_Data.xlsx"  
    sheet_name = "Invalid Accounts"  

    try:
        workbook = openpyxl.load_workbook(excel_file_path)
    except FileNotFoundError:
        workbook = Workbook()

    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
    else:
        sheet = workbook.create_sheet(sheet_name)

        headers = ["ScomuserNum", "ID", "SocialMedia", "SocialMediaType", "ScomUserName", "URL", "Followers", "Validated", "ProxyJobID"]
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            sheet[f"{col_letter}1"] = header

    new_row = [
        user.scomuserNum,
        user.id,
        user.socialMedia,
        user.socialMediaType,
        user.scomUserName,
        user.url,
        user.followers,
        user.validated,
        user.proxyJobID
    ]
    sheet.append(new_row)

    
    workbook.save(excel_file_path)

def addValidUsersToExcelSheet(user):
    excel_file_path = "Daily_User_Data.xlsx"  
    sheet_name = "valid Accounts"  

    
    try:
        workbook = openpyxl.load_workbook(excel_file_path)
    except FileNotFoundError:
        workbook = Workbook()

    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
    else:
        sheet = workbook.create_sheet(sheet_name)
        
        headers = ["ScomuserNum", "ID", "SocialMedia", "SocialMediaType", "ScomUserName", "URL", "Followers", "Validated", "ProxyJobID"]
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            sheet[f"{col_letter}1"] = header

    new_row = [
        user.scomuserNum,
        user.id,
        user.socialMedia,
        user.socialMediaType,
        user.scomUserName,
        user.url,
        user.followers,
        user.validated,
        user.proxyJobID
    ]
    sheet.append(new_row)

    
    workbook.save(excel_file_path)


def addInvalidPostsToExcelSheet(post):
    excel_file_path = "Daily_User_Data.xlsx"  
    sheet_name = "Invalid Posts"  

    
    try:
        workbook = openpyxl.load_workbook(excel_file_path)
    except FileNotFoundError:
        workbook = Workbook()

    
    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
    else:
        sheet = workbook.create_sheet(sheet_name)

        
        headers = ["ScomuserNum", "ID", "SocialMedia", "SocialMediaType", "ScomUserName", "URL", "Followers", "Validated", "ProxyJobID", "Reward", "Wallet Addy"]
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            sheet[f"{col_letter}1"] = header

    
    new_row = [
        post.scomuserNum,
        post.id,
        post.socialMedia,
        post.socialMediaType,
        post.scomUserName,
        post.url,
        post.followers,
        post.validated,
        post.proxyJobID,
        post.totalPayment,
        post.walletAddr
    ]
    sheet.append(new_row)
    workbook.save(excel_file_path)


def addValidPostsToExcelSheet(post):
    excel_file_path = "Daily_User_Data.xlsx"  
    sheet_name = "Valid Posts To Be Rewarded"  

    
    try:
        workbook = openpyxl.load_workbook(excel_file_path)
    except FileNotFoundError:
        workbook = Workbook()

    
    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
    else:
        sheet = workbook.create_sheet(sheet_name)

        
        headers = ["ScomuserNum", "ID", "SocialMedia", "SocialMediaType", "ScomUserName", "URL", "Followers", "Validated", "ProxyJobID", "Reward", "Wallet Addy"]
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            sheet[f"{col_letter}1"] = header

    
    new_row = [
        post.scomuserNum,
        post.id,
        post.socialMedia,
        post.socialMediaType,
        post.scomUserName,
        post.url,
        post.followers,
        post.validated,
        post.proxyJobID,
        post.totalPayment,
        post.walletAddr
    ]
    sheet.append(new_row)
    workbook.save(excel_file_path)


def add_InValid_Followers_But_Valid_Posts_To_Excel_Sheet(post):
    excel_file_path = "Daily_User_Data.xlsx"  
    sheet_name = "Valid Posts -1 Followers" 

    
    try:
        workbook = openpyxl.load_workbook(excel_file_path)
    except FileNotFoundError:
        workbook = Workbook()

    
    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
    else:
        sheet = workbook.create_sheet(sheet_name)

        
        headers = ["ScomuserNum", "ID", "SocialMedia", "SocialMediaType", "ScomUserName", "URL", "Followers", "Validated", "ProxyJobID", "Reward", "Wallet Addy"]
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            sheet[f"{col_letter}1"] = header

    
    new_row = [
        post.scomuserNum,
        post.id,
        post.socialMedia,
        post.socialMediaType,
        post.scomUserName,
        post.url,
        post.followers,
        post.validated,
        post.proxyJobID,
        post.totalPayment,
        post.walletAddr
    ]
    sheet.append(new_row)
    workbook.save(excel_file_path)


def addInstaStoryToExcelSheet(post):
    excel_file_path = "Daily_User_Data.xlsx"  
    sheet_name = "Instagram Stories" 

    
    try:
        workbook = openpyxl.load_workbook(excel_file_path)
    except FileNotFoundError:
        workbook = Workbook()

    
    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
    else:
        sheet = workbook.create_sheet(sheet_name)

        
        headers = ["ScomuserNum", "ID", "SocialMedia", "SocialMediaType", "ScomUserName", "URL", "Followers", "Validated", "ProxyJobID", "Reward", "Wallet Addy"]
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            sheet[f"{col_letter}1"] = header

    
    new_row = [
        post.scomuserNum,
        post.id,
        post.socialMedia,
        post.socialMediaType,
        post.scomUserName,
        post.url,
        post.followers,
        post.validated,
        post.proxyJobID,
        post.totalPayment,
        post.walletAddr
    ]
    sheet.append(new_row)
    workbook.save(excel_file_path)

